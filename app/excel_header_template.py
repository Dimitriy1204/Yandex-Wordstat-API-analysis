#!/usr/bin/env python3
"""Заголовки таблиц Excel: из образца yandex_analysis_exemple.xlsx или встроенные."""
from __future__ import annotations
import os
import sys
import logging
from typing import Any

logger = logging.getLogger("wordstat.excel_headers")

TEMPLATE_NAMES = ("yandex_analysis_exemple.xlsx", "yandex_analysis_example.xlsx")

# (текст, цвет или None) — для write_rich_string в xlsxwriter
HeaderRun = tuple[str, str | None]
HeaderRuns = list[HeaderRun]

DEFAULT_HEADERS: dict[str, list[str]] = {
    "Данные": [
        "Дата\nмесяц",
        "Поисковый запрос",
        "Показы в месяц",
        "Скользящее среднее\nза 3 мес.",
        "Темп роста, %",
        "Волатильность\n🟢низкая 🟡средняя 🔴высокая",
        "Индекс сезонности",
        "Индекс конкуренции\n🟢<0.40 🟡0.40-0.60 🔴>0.60",
    ],
    "Сводка по запросам": [
        "Поисковый запрос",
        "Средняя частотность",
        "Суммарная частотность\n(ёмкость рынка)",
        "Стандартное отклонение",
        "Средний рост, %\n(за весь период)",
        "Импульс роста, %\n(последние 6 мес.)",
        "Средний рост, %\n(последний год)",
        "Наклон тренда\n(показов/мес)",
        "Волатильность\n🟢низкая 🟡средняя 🔴высокая",
        "Индекс конкуренции\n🟢<0.40 🟡0.40-0.60 🔴>0.60",
        "Прогноз сред.\nпоказов/мес\n(6 мес.)",
    ],
    "Прогноз": [
        "Поисковый запрос",
        "Дата\nмесяц",
        "Прогноз показов",
        "Нижняя граница",
        "Верхняя граница",
        "Примечание",
    ],
    "ocean": [
        "№",
        "Поисковый запрос",
        "Индекс конкуренции\n🟢<0.40 🟡0.40-0.60 🔴>0.60",
        "Средний рост за последние 6 мес. (%)\n(значение метрики)",
        "Суммарная частотность\nЕмкость рынка",
        "Вердикт",
        "Вердикт нейросети\n(YandexGPT)",
    ],
    "growing": [
        "№",
        "Поисковый запрос",
        "Средний рост за последние 6 мес. (%)\n(значение метрики)",
        "Суммарная частотность\nЕмкость рынка",
        "Индекс конкуренции\n🟢<0.40 🟡0.40-0.60 🔴>0.60",
        "Волатильность\n🟢низкая 🟡средняя 🔴высокая",
        "Вердикт",
    ],
}

_GROWTH_HDR_OVERRIDE = {
    "growth_momentum": "Средний рост за последние 6 мес. (%)\n(значение метрики)",
    "mean_growth": "Средний рост за весь период (%)\n(значение метрики)",
    "trend_slope": "Наклон линейного тренда\n(показов/мес)",
}

# Индекс колонки метрики роста (0-based) в таблицах ocean / growing
_GROWTH_COL_INDEX = {"ocean": 3, "growing": 2}

_cached_headers: dict[str, list[str]] | None = None
_cached_runs: dict[str, list[HeaderRuns]] | None = None


def _search_dirs() -> list[str]:
    dirs: list[str] = []
    if getattr(sys, "frozen", False):
        # Сначала встроенный в exe образец (_MEIPASS), затем файл рядом с exe
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            dirs.append(meipass)
        dirs.append(os.path.dirname(os.path.abspath(sys.executable)))
    here = os.path.dirname(os.path.abspath(__file__))
    dirs.append(os.path.dirname(here))
    dirs.append(os.getcwd())
    return dirs


def find_template_path() -> str | None:
    for d in _search_dirs():
        for name in TEMPLATE_NAMES:
            path = os.path.join(d, name)
            if os.path.isfile(path):
                return path
    return None


def _rgb_to_xlsx(color: Any) -> str | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None) or str(color)
    if not rgb or rgb in ("00000000", "0", "values must be of type"):
        return None
    rgb = str(rgb)
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    return "#" + rgb.upper()


def _cell_runs(cell) -> HeaderRuns:
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    val = cell.value
    if val is None:
        return []
    if isinstance(val, CellRichText):
        runs: HeaderRuns = []
        for item in val:
            if isinstance(item, TextBlock):
                color = _rgb_to_xlsx(item.font.color if item.font else None)
                runs.append((item.text or "", color))
            elif item:
                runs.append((str(item), None))
        return runs
    return [(str(val), None)]


def _runs_to_text(runs: HeaderRuns) -> str:
    return "".join(t for t, _ in runs)


def _load_from_xlsx(path: str) -> tuple[dict[str, list[str]], dict[str, list[HeaderRuns]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=False, data_only=False)
    headers_out: dict[str, list[str]] = {}
    runs_out: dict[str, list[HeaderRuns]] = {}

    sheet_rows = {
        "Данные": 2,
        "Сводка по запросам": 2,
        "Прогноз": 2,
        "Голубые океаны": 2,
        "Растущие рынки": 2,
    }
    for sheet_name, row_idx in sheet_rows.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        texts: list[str] = []
        col_runs: list[HeaderRuns] = []
        col = 1
        while col <= 32:
            cell = ws.cell(row=row_idx, column=col)
            runs = _cell_runs(cell)
            text = _runs_to_text(runs).strip()
            if not text:
                break
            texts.append(text)
            col_runs.append(runs if runs else [(text, None)])
            col += 1
        if texts:
            key = "ocean" if sheet_name == "Голубые океаны" else (
                "growing" if sheet_name == "Растущие рынки" else sheet_name
            )
            headers_out[key] = texts
            runs_out[key] = col_runs
            logger.info("Заголовки из образца «%s»: %s колонок", sheet_name, len(texts))
    wb.close()
    return headers_out, runs_out


def _ensure_loaded() -> None:
    global _cached_headers, _cached_runs
    if _cached_headers is not None:
        return
    tpl = find_template_path()
    if tpl:
        try:
            h, r = _load_from_xlsx(tpl)
            _cached_headers = h
            _cached_runs = r
            logger.info("Шаблон заголовков: %s", tpl)
        except Exception as e:
            logger.warning("Не удалось прочитать %s: %s", tpl, e)
            _cached_headers = {}
            _cached_runs = {}
    else:
        _cached_headers = {}
        _cached_runs = {}
        logger.info(
            "Образец Excel не найден (%s). Используются встроенные заголовки.",
            ", ".join(TEMPLATE_NAMES),
        )


def _normalize_ocean_growing_headers(headers: list[str]) -> list[str]:
    """Убирает лишнюю колонку «Прогноз» из образца, если структура старше текущей."""
    if len(headers) == 8 and "прогноз" in headers[5].lower():
        return headers[:5] + headers[6:]
    return headers


def _apply_growth_override(
    sheet_key: str,
    headers: list[str],
    runs: list[HeaderRuns] | None,
    growth_col: str | None,
) -> tuple[list[str], list[HeaderRuns] | None]:
    if not growth_col or growth_col not in _GROWTH_HDR_OVERRIDE:
        return headers, runs
    idx = _GROWTH_COL_INDEX.get(sheet_key)
    if idx is None or len(headers) <= idx:
        return headers, runs
    new_text = _GROWTH_HDR_OVERRIDE[growth_col]
    headers = list(headers)
    headers[idx] = new_text
    if runs and len(runs) > idx:
        runs = list(runs)
        runs[idx] = [(new_text, None)]
    return headers, runs


def get_table_headers(sheet_key: str, growth_col: str | None = None) -> list[str]:
    _ensure_loaded()
    if _cached_headers and sheet_key in _cached_headers:
        h = _normalize_ocean_growing_headers(list(_cached_headers[sheet_key]))
    else:
        h = list(DEFAULT_HEADERS[sheet_key])
    h, _ = _apply_growth_override(sheet_key, h, None, growth_col)
    return h


def get_table_header_runs(sheet_key: str, growth_col: str | None = None) -> list[HeaderRuns] | None:
    """Rich-text фрагменты из образца; None — белый текст заголовка без раскраски эмодзи."""
    _ensure_loaded()
    if not _cached_runs or sheet_key not in _cached_runs:
        return None
    runs = [list(col) for col in _cached_runs[sheet_key]]
    texts = _normalize_ocean_growing_headers([_runs_to_text(c) for c in runs])
    if len(texts) != len(runs):
        runs = [[(t, None)] for t in texts]
    _, runs = _apply_growth_override(sheet_key, texts, runs, growth_col)
    return runs
